"""
Leitura e interpretacao das planilhas LPU (Lista de Precos Unitarios).
"""
import re
import openpyxl


def detectar_disciplina(nome_arquivo: str) -> str:
    nome = nome_arquivo.lower()
    if "civil" in nome:
        return "Civil"
    if "eletric" in nome or "elétric" in nome:
        return "Eletrica"
    return "Nao identificada"


def _achar_valor(ws, rotulo: str):
    """Procura por uma celula cujo texto comeca com `rotulo` na aba Capa
    e retorna o valor da celula vizinha (a direita)."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower().startswith(rotulo.lower()):
                vizinho = ws.cell(row=cell.row, column=cell.column + 1)
                if vizinho.value not in (None, ""):
                    return vizinho.value
    return None


def ler_dados_capa(caminho_arquivo: str) -> dict:
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    if "Capa" not in wb.sheetnames:
        return {}
    ws = wb["Capa"]
    codigo = _achar_valor(ws, "Sorting Code")
    local = _achar_valor(ws, "Site")
    endereco = _achar_valor(ws, "Endereço") or _achar_valor(ws, "Endereco")
    revisao = _achar_valor(ws, "Revisão orçamento") or _achar_valor(ws, "Revisao orcamento")
    prazo_execucao = _achar_valor(ws, "Prazo estimado de execução") or _achar_valor(
        ws, "Prazo estimado de execucao"
    )
    return {
        "codigo_projeto": (str(codigo).strip() if codigo else ""),
        "local": (str(local).strip() if local else ""),
        "endereco": (str(endereco).strip() if endereco else ""),
        "revisao": (str(revisao).strip() if revisao else ""),
        "prazo_execucao": (str(prazo_execucao).strip() if prazo_execucao else ""),
    }


def _achar_aba_orcamentaria(wb):
    for nome in wb.sheetnames:
        if "orçament" in nome.lower() or "orcament" in nome.lower():
            return nome
    return None


def ler_itens_orcamento(caminho_arquivo: str) -> list:
    """Retorna apenas os itens-folha com quantidade preenchida (> 0),
    que sao os que efetivamente serao executados."""
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    nome_aba = _achar_aba_orcamentaria(wb)
    if not nome_aba:
        return []
    ws = wb[nome_aba]

    header_row = 1
    headers = [c.value for c in ws[header_row]]
    col = {h: idx + 1 for idx, h in enumerate(headers) if h}

    col_codigo = col.get("Código") or col.get("Codigo")
    col_item = col.get("Item")
    col_qtd = col.get("Qtd.") or col.get("Qtd")
    col_unid = col.get("Unid.") or col.get("Unid")
    col_obs = col.get("Observações de RFP") or col.get("Observacoes de RFP")

    itens = []
    for row in ws.iter_rows(min_row=header_row + 1):
        qtd_cell = row[col_qtd - 1] if col_qtd else None
        qtd = qtd_cell.value if qtd_cell else None
        if not isinstance(qtd, (int, float)) or qtd <= 0:
            continue

        codigo = row[col_codigo - 1].value if col_codigo else ""
        descricao = row[col_item - 1].value if col_item else ""
        unidade = row[col_unid - 1].value if col_unid else ""
        obs = row[col_obs - 1].value if col_obs else ""

        itens.append(
            {
                "codigo": str(codigo) if codigo else "",
                "descricao": str(descricao).strip() if descricao else "",
                "quantidade": qtd,
                "unidade": str(unidade).strip() if unidade else "",
                "observacao_rfp": str(obs).strip() if obs else "",
            }
        )
    return itens


def ler_valor_total_bdi(caminho_arquivo: str):
    """Retorna o valor total do orcamento (com BDI), somando as linhas de
    Nivel 1 (categorias-raiz) da aba orcamentaria."""
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    nome_aba = _achar_aba_orcamentaria(wb)
    if not nome_aba:
        return None
    ws = wb[nome_aba]

    headers = [c.value for c in ws[1]]
    col = {h: idx + 1 for idx, h in enumerate(headers) if h}
    col_nivel = col.get("Nível") or col.get("Nivel")
    col_bdi = col.get("Valor total + BDI")
    if not col_nivel or not col_bdi:
        return None

    total = 0.0
    encontrou = False
    for row in ws.iter_rows(min_row=2):
        nivel = row[col_nivel - 1].value
        if nivel == 1:
            valor = row[col_bdi - 1].value
            if isinstance(valor, (int, float)):
                total += valor
                encontrou = True
    return total if encontrou else None


def carregar_lpu(caminho_arquivo: str, nome_arquivo: str) -> dict:
    dados_capa = ler_dados_capa(caminho_arquivo)
    itens = ler_itens_orcamento(caminho_arquivo)
    disciplina = detectar_disciplina(nome_arquivo)
    valor_total_bdi = ler_valor_total_bdi(caminho_arquivo)
    return {
        **dados_capa,
        "disciplina": disciplina,
        "itens": itens,
        "valor_total_bdi": valor_total_bdi,
    }
